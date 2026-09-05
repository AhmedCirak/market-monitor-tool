"""
Generisanje Excel reporta.

Princip: sve brojke racuna Pandas i upisuju se kao vrijednosti.
Excel formule postoje samo na sheetu "Market Overview" - tako je dashboard
ziv (reaguje na izmjene i filtriranje), a tesku matematiku ne radi Excel.

Svi data sheetovi su prave Excel Tabele (ListObject), pa se automatski
prosiruju kad skripta doda nove redove i mogu se odmah koristiti kao
izvor za pivot tabele.
"""

from __future__ import annotations

import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import config

logger = logging.getLogger(__name__)

# --------------------------------------------------------------------------
# Stil
# --------------------------------------------------------------------------
FONT = "Arial"
F_TITLE = Font(name=FONT, size=16, bold=True, color="1F3864")
F_SECTION = Font(name=FONT, size=11, bold=True, color="FFFFFF")
F_LABEL = Font(name=FONT, size=10)
F_VALUE = Font(name=FONT, size=10, bold=True)
F_INPUT = Font(name=FONT, size=10, bold=True, color="0000FF")   # plavo = rucni unos
F_NOTE = Font(name=FONT, size=9, italic=True, color="808080")
F_HEADER = Font(name=FONT, size=10, bold=True, color="FFFFFF")

FILL_SECTION = PatternFill("solid", fgColor="1F3864")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")
FILL_CARD = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Formati brojeva
FMT_PCT = "0.00%;[Red]-0.00%;-"
FMT_PRICE = "#,##0.00"
FMT_PRICE_FX = "#,##0.0000"
FMT_INT = "#,##0"
FMT_RATIO = '0.00"x"'
FMT_NUM1 = "0.0"
FMT_NUM2 = "0.00"
FMT_DATE = "DD.MM.YYYY"

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
GREEN_FONT = Font(name=FONT, size=10, color="006100")
RED_FONT = Font(name=FONT, size=10, color="9C0006")


# --------------------------------------------------------------------------
# Pomocne funkcije
# --------------------------------------------------------------------------
def _write_table(ws, df: pd.DataFrame, table_name: str, formats: dict[str, str],
                 start_row: int = 1) -> str | None:
    """Upise DataFrame kao pravu Excel Tabelu i vrati njen raspon."""
    if df.empty:
        ws.cell(start_row, 1, "Nema podataka.").font = F_NOTE
        return None

    # Zaglavlje
    for col_idx, column in enumerate(df.columns, start=1):
        cell = ws.cell(start_row, col_idx, column)
        cell.font = F_HEADER
        cell.fill = FILL_SECTION
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # Podaci
    for row_offset, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for col_idx, column in enumerate(df.columns, start=1):
            value = row[column]
            if pd.isna(value):
                value = None
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            elif hasattr(value, "item"):
                value = value.item()

            cell = ws.cell(row_offset, col_idx, value)
            cell.font = F_LABEL
            cell.border = BORDER
            if column in formats:
                cell.number_format = formats[column]

    last_row = start_row + len(df)
    last_col = get_column_letter(len(df.columns))
    ref = f"A{start_row}:{last_col}{last_row}"

    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)

    # Sirine kolona
    for col_idx, column in enumerate(df.columns, start=1):
        width = max(len(str(column)) + 4, 12)
        if column in ("Instrument", "Opis"):
            width = 34
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = ws.cell(start_row + 1, 1)
    return ref


def _color_scale_on(ws, df: pd.DataFrame, column: str, start_row: int = 1) -> None:
    """Zeleno-crvena skala na jednoj koloni tabele."""
    if df.empty or column not in df.columns:
        return
    idx = list(df.columns).index(column) + 1
    letter = get_column_letter(idx)
    rng = f"{letter}{start_row + 1}:{letter}{start_row + len(df)}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="max", end_color="63BE7B",
    ))


def _updown_format_on(ws, df: pd.DataFrame, column: str, start_row: int = 1) -> None:
    """Zeleno za pozitivno, crveno za negativno."""
    if df.empty or column not in df.columns:
        return
    idx = list(df.columns).index(column) + 1
    letter = get_column_letter(idx)
    rng = f"{letter}{start_row + 1}:{letter}{start_row + len(df)}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThan", formula=["0"], font=GREEN_FONT))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=["0"], font=RED_FONT))


def _section(ws, row: int, title: str, width: int = 6) -> int:
    """Naslov sekcije preko vise kolona. Vrati sljedeci slobodan red."""
    for col in range(1, width + 1):
        cell = ws.cell(row, col)
        cell.fill = FILL_SECTION
        if col == 1:
            cell.value = title
            cell.font = F_SECTION
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 20
    return row + 2


# --------------------------------------------------------------------------
# Priprema podataka za sheetove
# --------------------------------------------------------------------------
CATEGORY_COLUMNS = {
    "symbol": "Symbol",
    "name": "Instrument",
    "date": "Datum",
    "open": "Open",
    "high": "High",
    "low": "Low",
    "close": "Close",
    "return_1d": "Change_1D",
    "return_5d": "Change_5D",
    "return_20d": "Change_20D",
    "return_ytd": "YTD",
    "return_1y": "Return_1Y",
    "volume": "Volume",
    "volume_ratio": "Volume_Ratio",
    "high_52w": "High_52W",
    "low_52w": "Low_52W",
    "pct_from_52w_high": "From_52W_High",
    "volatility": "Volatilnost",
    "rsi": "RSI",
    "trend": "Trend",
}

TECHNICAL_COLUMNS = {
    "symbol": "Symbol",
    "name": "Instrument",
    "category": "Kategorija",
    "date": "Datum",
    "close": "Close",
    "return_1d": "Change_1D",
    "ma_20": "MA20",
    "ma_50": "MA50",
    "ma_200": "MA200",
    "price_vs_ma200": "Price_vs_MA200",
    "trend": "Trend",
    "rsi": "RSI",
    "macd_hist": "MACD_Hist",
    "volatility": "Volatilnost",
    "volume_ratio": "Volume_Ratio",
    "drawdown": "Drawdown",
    "pct_from_52w_high": "From_52W_High",
}

CATEGORY_FORMATS = {
    "Datum": FMT_DATE, "Open": FMT_PRICE, "High": FMT_PRICE, "Low": FMT_PRICE,
    "Close": FMT_PRICE, "Change_1D": FMT_PCT, "Change_5D": FMT_PCT,
    "Change_20D": FMT_PCT, "YTD": FMT_PCT, "Return_1Y": FMT_PCT,
    "Volume": FMT_INT, "Volume_Ratio": FMT_RATIO, "High_52W": FMT_PRICE,
    "Low_52W": FMT_PRICE, "From_52W_High": FMT_PCT, "Volatilnost": FMT_PCT,
    "RSI": FMT_NUM1,
}

TECHNICAL_FORMATS = {
    "Datum": FMT_DATE, "Close": FMT_PRICE, "Change_1D": FMT_PCT,
    "MA20": FMT_PRICE, "MA50": FMT_PRICE, "MA200": FMT_PRICE,
    "Price_vs_MA200": FMT_PCT, "RSI": FMT_NUM1, "MACD_Hist": FMT_NUM2,
    "Volatilnost": FMT_PCT, "Volume_Ratio": FMT_RATIO, "Drawdown": FMT_PCT,
    "From_52W_High": FMT_PCT, "Beta": FMT_NUM2,
}


def _prepare(snapshot: pd.DataFrame, mapping: dict, fx: bool = False) -> pd.DataFrame:
    cols = [c for c in mapping if c in snapshot.columns]
    df = snapshot[cols].rename(columns=mapping).copy()
    return df


# --------------------------------------------------------------------------
# Glavna funkcija
# --------------------------------------------------------------------------
def build_report(
    snapshot: pd.DataFrame,
    enriched: pd.DataFrame,
    alerts_df: pd.DataFrame,
    correlation: pd.DataFrame,
    betas: pd.Series,
    output_path=None,
) -> str:
    output_path = output_path or config.REPORT_FILE
    wb = Workbook()

    # Redoslijed sheetova
    ws_overview = wb.active
    ws_overview.title = "Market Overview"

    # ---------------- Kategorijski sheetovi ----------------
    table_names = {"Stocks": "tbl_Stocks", "FX": "tbl_FX", "Commodities": "tbl_Commodities"}
    for category, table_name in table_names.items():
        ws = wb.create_sheet(category)
        subset = snapshot[snapshot["category"] == category]
        df = _prepare(subset, CATEGORY_COLUMNS)
        formats = dict(CATEGORY_FORMATS)
        if category == "FX":
            for col in ("Open", "High", "Low", "Close", "High_52W", "Low_52W"):
                formats[col] = FMT_PRICE_FX
        _write_table(ws, df, table_name, formats)
        _updown_format_on(ws, df, "Change_1D")
        _color_scale_on(ws, df, "YTD")

    # ---------------- Technical Analysis ----------------
    ws_tech = wb.create_sheet("Technical Analysis")
    tech = _prepare(snapshot, TECHNICAL_COLUMNS)
    if not tech.empty and not betas.empty:
        tech["Beta"] = tech["Symbol"].map(betas)
    _write_table(ws_tech, tech, "tbl_Technical", TECHNICAL_FORMATS)
    _updown_format_on(ws_tech, tech, "Change_1D")
    _color_scale_on(ws_tech, tech, "RSI")

    # ---------------- Alerts ----------------
    ws_alerts = wb.create_sheet("Alerts")
    alerts_out = alerts_df.copy()
    if not alerts_out.empty:
        alerts_out["Datum"] = pd.to_datetime(alerts_out["Datum"])
    _write_table(ws_alerts, alerts_out, "tbl_Alerts",
                 {"Datum": FMT_DATE, "Vrijednost": FMT_NUM2})

    # ---------------- Correlation ----------------
    ws_corr = wb.create_sheet("Correlation")
    _write_correlation(ws_corr, correlation)

    # ---------------- Price History + grafikon ----------------
    ws_hist = wb.create_sheet("Price History")
    _write_price_history(ws_hist, enriched)

    # ---------------- Pivot Data ----------------
    ws_pivot = wb.create_sheet("Pivot Data")
    _write_pivot_data(ws_pivot, enriched)

    # ---------------- Overview (zadnji, jer referencira ostale) ----------------
    _write_overview(ws_overview, snapshot, alerts_df)

    wb.save(output_path)
    logger.info("Report snimljen: %s", output_path)
    return str(output_path)


# --------------------------------------------------------------------------
# Market Overview - jedini sheet sa pravim Excel formulama
# --------------------------------------------------------------------------
def _write_overview(ws, snapshot: pd.DataFrame, alerts_df: pd.DataFrame) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    for col in "BCDE":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["F"].width = 20

    ws["A1"] = "MARKET OVERVIEW"
    ws["A1"].font = F_TITLE
    ws["A2"] = "Automatski generisan pregled stanja trzista"
    ws["A2"].font = F_NOTE

    last_date = pd.to_datetime(snapshot["date"]).max() if not snapshot.empty else None
    ws["A3"] = "Zadnji podatak (datum bara):"
    ws["A3"].font = F_LABEL
    ws["B3"] = last_date.to_pydatetime() if last_date is not None else "-"
    ws["B3"].font = F_VALUE
    ws["B3"].number_format = FMT_DATE

    ws["A4"] = "Report generisan:"
    ws["A4"].font = F_LABEL
    ws["B4"] = datetime.now()
    ws["B4"].font = F_VALUE
    ws["B4"].number_format = "DD.MM.YYYY HH:MM"

    row = 6

    # ---- Parametri (plavo = rucni unos, formule ih referenciraju) ----
    row = _section(ws, row, "PARAMETRI  (plave celije mozes mijenjati - formule ispod se same preracunaju)")
    params = [
        ("Prag dnevne promjene", config.ALERT_PRICE_CHANGE_PCT / 100, FMT_PCT, "prag_promjena"),
        ("RSI overbought", config.ALERT_RSI_OVERBOUGHT, FMT_NUM1, "prag_rsi_high"),
        ("RSI oversold", config.ALERT_RSI_OVERSOLD, FMT_NUM1, "prag_rsi_low"),
        ("Volume multiplikator", config.ALERT_VOLUME_MULTIPLIER, FMT_RATIO, "prag_volume"),
    ]
    param_cells = {}
    for label, value, fmt, key in params:
        ws.cell(row, 1, label).font = F_LABEL
        cell = ws.cell(row, 2, value)
        cell.font = F_INPUT
        cell.fill = FILL_INPUT
        cell.number_format = fmt
        cell.border = BORDER
        param_cells[key] = f"$B${row}"
        row += 1
    row += 1

    # ---- Sazetak trzista ----
    row = _section(ws, row, "SAZETAK TRZISTA")
    total = "COUNTA(tbl_Technical[Instrument])"
    summary = [
        ("Ukupno pracenih instrumenata", f"={total}", FMT_INT),
        ("Instrumenata u rastu", '=COUNTIF(tbl_Technical[Change_1D],">0")', FMT_INT),
        ("Instrumenata u padu", '=COUNTIF(tbl_Technical[Change_1D],"<0")', FMT_INT),
        ("Prosjecna dnevna promjena", "=IFERROR(AVERAGE(tbl_Technical[Change_1D]),0)", FMT_PCT),
        ("Medijana dnevne promjene", "=IFERROR(MEDIAN(tbl_Technical[Change_1D]),0)", FMT_PCT),
        ("Najveci pojedinacni rast", "=IFERROR(MAX(tbl_Technical[Change_1D]),0)", FMT_PCT),
        ("Najveci pojedinacni pad", "=IFERROR(MIN(tbl_Technical[Change_1D]),0)", FMT_PCT),
        ("Iznad 200-dnevnog prosjeka",
         '=COUNTIF(tbl_Technical[Price_vs_MA200],">0")', FMT_INT),
        ("U uzlaznom trendu (MA50 > MA200)",
         '=COUNTIF(tbl_Technical[Trend],"Uzlazni")', FMT_INT),
    ]
    for label, formula, fmt in summary:
        ws.cell(row, 1, label).font = F_LABEL
        cell = ws.cell(row, 2, formula)
        cell.font = F_VALUE
        cell.number_format = fmt
        cell.fill = FILL_CARD
        cell.border = BORDER
        row += 1
    row += 1

    # ---- Alerti ----
    row = _section(ws, row, "ALERTI")
    alert_rows = [
        ("Ukupno alerta", "=COUNTA(tbl_Alerts[Instrument])", FMT_INT),
        ("Instrumenata sa bar jednim alertom",
         "=SUMPRODUCT(--(COUNTIF(tbl_Alerts[Instrument],tbl_Technical[Instrument])>0))", FMT_INT),
        ("Udio instrumenata u alert zoni",
         "=IFERROR(SUMPRODUCT(--(COUNTIF(tbl_Alerts[Instrument],tbl_Technical[Instrument])>0))"
         f"/{total},0)", FMT_PCT),
        ("RSI overbought", f'=COUNTIF(tbl_Technical[RSI],">="&{param_cells["prag_rsi_high"]})', FMT_INT),
        ("RSI oversold", f'=COUNTIF(tbl_Technical[RSI],"<="&{param_cells["prag_rsi_low"]})', FMT_INT),
        ("Volume anomalije",
         f'=COUNTIF(tbl_Technical[Volume_Ratio],">="&{param_cells["prag_volume"]})', FMT_INT),
        ("Promjena iznad praga",
         f'=COUNTIF(tbl_Technical[Change_1D],">="&{param_cells["prag_promjena"]})'
         f'+COUNTIF(tbl_Technical[Change_1D],"<="&-{param_cells["prag_promjena"]})', FMT_INT),
    ]
    for label, formula, fmt in alert_rows:
        ws.cell(row, 1, label).font = F_LABEL
        cell = ws.cell(row, 2, formula)
        cell.font = F_VALUE
        cell.number_format = fmt
        cell.fill = FILL_CARD
        cell.border = BORDER
        row += 1
    row += 1

    # ---- Po kategorijama ----
    row = _section(ws, row, "PO KATEGORIJAMA")
    headers = ["Kategorija", "Broj", "U rastu", "U padu", "Prosj. promjena"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        cell.font = F_HEADER
        cell.fill = FILL_SECTION
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for category in ("Stocks", "FX", "Commodities"):
        ws.cell(row, 1, category).font = F_LABEL
        ws.cell(row, 1).border = BORDER
        cat_ref = f"$A${row}"
        cells = [
            (2, f"=COUNTIF(tbl_Technical[Kategorija],{cat_ref})", FMT_INT),
            (3, f'=COUNTIFS(tbl_Technical[Kategorija],{cat_ref},tbl_Technical[Change_1D],">0")', FMT_INT),
            (4, f'=COUNTIFS(tbl_Technical[Kategorija],{cat_ref},tbl_Technical[Change_1D],"<0")', FMT_INT),
            (5, f"=IFERROR(AVERAGEIF(tbl_Technical[Kategorija],{cat_ref},tbl_Technical[Change_1D]),0)", FMT_PCT),
        ]
        for col, formula, fmt in cells:
            cell = ws.cell(row, col, formula)
            cell.font = F_LABEL
            cell.number_format = fmt
            cell.border = BORDER
        row += 1
    row += 1

    # ---- Top 5 dobitnika / gubitnika ----
    row = _section(ws, row, "TOP 5 NAJJACIH I NAJSLABIJIH (dnevna promjena)")
    ws.cell(row, 1, "Dobitnik").font = F_HEADER
    ws.cell(row, 2, "Promjena").font = F_HEADER
    ws.cell(row, 4, "Najslabiji").font = F_HEADER
    ws.cell(row, 5, "Promjena").font = F_HEADER
    for col in (1, 2, 4, 5):
        ws.cell(row, col).fill = FILL_SECTION
        ws.cell(row, col).border = BORDER
    row += 1

    for rank in range(1, 6):
        large = f"LARGE(tbl_Technical[Change_1D],{rank})"
        small = f"SMALL(tbl_Technical[Change_1D],{rank})"

        c = ws.cell(row, 1, f"=IFERROR(INDEX(tbl_Technical[Instrument],"
                            f"MATCH({large},tbl_Technical[Change_1D],0)),\"-\")")
        c.font = F_LABEL
        c.border = BORDER
        c = ws.cell(row, 2, f"=IFERROR({large},\"-\")")
        c.font = F_VALUE
        c.number_format = FMT_PCT
        c.border = BORDER

        c = ws.cell(row, 4, f"=IFERROR(INDEX(tbl_Technical[Instrument],"
                            f"MATCH({small},tbl_Technical[Change_1D],0)),\"-\")")
        c.font = F_LABEL
        c.border = BORDER
        c = ws.cell(row, 5, f"=IFERROR({small},\"-\")")
        c.font = F_VALUE
        c.number_format = FMT_PCT
        c.border = BORDER
        row += 1
    row += 1

    # ---- Alerti po tipu ----
    row = _section(ws, row, "ALERTI PO TIPU")
    alert_types = [
        "Promjena cijene", "RSI overbought", "RSI oversold", "Volume anomalija",
        "Blizu 52w maksimuma", "Blizu 52w minimuma", "Drawdown", "Test MA200",
    ]
    for alert_type in alert_types:
        ws.cell(row, 1, alert_type).font = F_LABEL
        ws.cell(row, 1).border = BORDER
        cell = ws.cell(row, 2, f'=COUNTIF(tbl_Alerts[Tip],$A${row})')
        cell.font = F_VALUE
        cell.number_format = FMT_INT
        cell.border = BORDER
        row += 1
    row += 1

    ws.cell(row, 1,
            "Napomena: sve brojke iznad su Excel formule nad tabelama sa ostalih sheetova. "
            "Kada skripta doda nove redove, tabele se prosire i ove formule se same azuriraju."
            ).font = F_NOTE
    row += 1
    ws.cell(row, 1,
            "Izvor podataka: Stooq (stooq.com), dnevni OHLCV. Pokazatelje racuna Pandas."
            ).font = F_NOTE


# --------------------------------------------------------------------------
# Ostali sheetovi
# --------------------------------------------------------------------------
def _write_correlation(ws, correlation: pd.DataFrame) -> None:
    ws["A1"] = "KORELACIJA DNEVNIH PRINOSA"
    ws["A1"].font = F_TITLE
    ws["A2"] = f"Zadnjih {config.CORRELATION_WINDOW} trgovackih dana. " \
               "1,0 = krecu se identicno, 0 = nepovezano, -1,0 = suprotno."
    ws["A2"].font = F_NOTE

    if correlation.empty:
        ws["A4"] = "Nema dovoljno podataka."
        ws["A4"].font = F_NOTE
        return

    start = 4
    ws.column_dimensions["A"].width = 24
    for col_idx, name in enumerate(correlation.columns, start=2):
        cell = ws.cell(start, col_idx, name)
        cell.font = F_HEADER
        cell.fill = FILL_SECTION
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True, textRotation=45)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    for row_idx, (name, values) in enumerate(correlation.iterrows(), start=start + 1):
        cell = ws.cell(row_idx, 1, name)
        cell.font = F_VALUE
        cell.border = BORDER
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(row_idx, col_idx, None if pd.isna(value) else float(value))
            cell.font = F_LABEL
            cell.number_format = FMT_NUM2
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

    last_row = start + len(correlation)
    last_col = get_column_letter(1 + len(correlation.columns))
    ws.conditional_formatting.add(
        f"B{start + 1}:{last_col}{last_row}",
        ColorScaleRule(
            start_type="num", start_value=-1, start_color="F8696B",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="num", end_value=1, end_color="63BE7B",
        ),
    )
    ws.freeze_panes = ws.cell(start + 1, 2)


def _write_price_history(ws, enriched: pd.DataFrame) -> None:
    from analytics import normalized_history_wide

    ws["A1"] = "KRETANJE CIJENA (normalizovano, prvi dan = 100)"
    ws["A1"].font = F_TITLE
    ws["A2"] = "Normalizacija omogucava poredjenje instrumenata razlicitih cjenovnih nivoa."
    ws["A2"].font = F_NOTE

    wide = normalized_history_wide(enriched)
    if wide.empty:
        ws["A4"] = "Nema dovoljno podataka."
        ws["A4"].font = F_NOTE
        return

    start = 4
    ws.cell(start, 1, "Datum").font = F_HEADER
    ws.cell(start, 1).fill = FILL_SECTION
    ws.column_dimensions["A"].width = 14

    for col_idx, name in enumerate(wide.columns, start=2):
        cell = ws.cell(start, col_idx, name)
        cell.font = F_HEADER
        cell.fill = FILL_SECTION
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    for row_idx, (date, values) in enumerate(wide.iterrows(), start=start + 1):
        cell = ws.cell(row_idx, 1, date.to_pydatetime())
        cell.number_format = FMT_DATE
        cell.font = F_LABEL
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(row_idx, col_idx, None if pd.isna(value) else float(value))
            cell.number_format = FMT_NUM1
            cell.font = F_LABEL

    last_row = start + len(wide)

    chart = LineChart()
    chart.title = "Relativno kretanje (prvi dan = 100)"
    chart.style = 2
    chart.y_axis.title = "Indeks"
    chart.x_axis.title = "Datum"
    chart.height = 12
    chart.width = 30

    data = Reference(ws, min_col=2, min_row=start,
                     max_col=1 + len(wide.columns), max_row=last_row)
    dates = Reference(ws, min_col=1, min_row=start + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(dates)
    for series in chart.series:
        series.smooth = False

    ws.add_chart(chart, f"{get_column_letter(len(wide.columns) + 3)}4")
    ws.freeze_panes = ws.cell(start + 1, 2)


def _write_pivot_data(ws, enriched: pd.DataFrame) -> None:
    """Long format - izvor za tvoje pivot tabele."""
    ws_note = (
        "Ovaj sheet je namjerno u 'long' formatu (jedan red = jedan instrument na jedan dan), "
        "sto je format koji pivot tabele ocekuju. Ubaci pivot preko Insert > PivotTable "
        "i kao izvor odaberi tabelu tbl_PivotData."
    )

    if enriched.empty:
        ws["A1"] = "Nema podataka."
        ws["A1"].font = F_NOTE
        return

    cutoff = enriched["date"].max() - pd.Timedelta(days=config.PIVOT_HISTORY_DAYS)
    subset = enriched[enriched["date"] >= cutoff].copy()

    df = pd.DataFrame({
        "Datum": subset["date"],
        "Godina": subset["date"].dt.year,
        "Mjesec": subset["date"].dt.to_period("M").astype(str),
        "Symbol": subset["symbol"],
        "Instrument": subset["name"],
        "Kategorija": subset["category"],
        "Close": subset["close"],
        "Volume": subset["volume"],
        "Change_1D": subset["return_1d"],
        "RSI": subset["rsi"],
        "Volatilnost": subset["volatility"],
    }).sort_values(["Instrument", "Datum"]).reset_index(drop=True)

    ws["A1"] = ws_note
    ws["A1"].font = F_NOTE
    ws.merge_cells("A1:H1")

    _write_table(ws, df, "tbl_PivotData", {
        "Datum": FMT_DATE, "Close": FMT_PRICE, "Volume": FMT_INT,
        "Change_1D": FMT_PCT, "RSI": FMT_NUM1, "Volatilnost": FMT_PCT,
    }, start_row=3)
